#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

#total 계산
row_num = 1
total = []
for row in ws:
	if row_num != 1:
		sum = ws.cell(row = row_num, column = 3).value * 0.3
		sum += ws.cell(row = row_num, column = 4).value * 0.35
		sum += ws.cell(row = row_num, column = 5).value * 0.34
		sum += ws.cell(row = row_num, column = 6).value
		ws.cell(row = row_num, column = 7).value = sum
		
		total.append(sum)
	row_num += 1

#grade 계산
sorted_total = sorted(total, reverse=True) #점수 내림차순 정렬
max_grade = [0, 0, 0, 0, 0] #몇 등까지 해당 학점(A+ ~ C+)을 받을 수 있는지
max_grade[1] = ((int)(len(total) // (10 / 3))) #A0
max_grade[0] = ((int)(max_grade[1] // 2)) #A+
max_grade[3] = ((int)(len(total) // (10 / 7))) #B0
max_grade[2] = ((int)((max_grade[3] - max_grade[1]) // 2 + max_grade[1])) #B+
max_grade[4] = ((int)((max_grade[3] + len(total)) // 2)) #C+

gradeNList = [1000, 1000, 1000, 1000, 1000]
index = 0
while index < 5:
	if max_grade[index] != 0:
		if max_grade[index] < len(total):
			if sorted_total[max_grade[index] - 1] == sorted_total[max_grade[index]]:
				i = max_grade[index] - 2
				if index == 0: endI = 0
				else: endI = max_grade[index - 1]
				
				while i > endI:
					if sorted_total[i] != sorted_total[max_grade[index]]:
						gradeNList[index] = sorted_total[i]
						break
					i -= 1
			else:
				gradeNList[index] = sorted_total[max_grade[index] - 1]
		elif max_grade[index] == len(total):
			gradeNList[index] = sorted_total[max_grade[index] - 1]
	index += 1

row_num2 = 1
for row in ws:
	if row_num2 != 1:
		if ws.cell(row = row_num2, column = 7).value >= gradeNList[0]:
			ws.cell(row = row_num2, column = 8).value = 'A+'
		elif ws.cell(row = row_num2, column = 7).value >= gradeNList[1]:
			ws.cell(row = row_num2, column = 8).value = 'A0'
		elif ws.cell(row = row_num2, column = 7).value >= gradeNList[2]:
			ws.cell(row = row_num2, column = 8).value = 'B+'
		elif ws.cell(row = row_num2, column = 7).value >= gradeNList[3]:
			ws.cell(row = row_num2, column = 8).value = 'B0'
		elif ws.cell(row = row_num2, column = 7).value >= gradeNList[4]:
			ws.cell(row = row_num2, column = 8).value = 'C+'
		else:
			ws.cell(row = row_num2, column = 8).value = 'C0'
	row_num2 += 1

wb.save("student.xlsx")
